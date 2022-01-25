from openpyxl import Workbook
from openpyxl.styles import Font

from .job_title import ExportJobTitle
from .job_grades import ExportJobGrade
from .nature_of_contract import ExportNatureOfContract
from .banks import ExportBankLists
# from .leave_category import ExportLeaveCategory
# from .leave_types import ExportLeaveTypes
from .holiday import ExportHoliday
from .employees import ExportEmployee


def generate_excel(fields, data):
    wb = Workbook()
    ws = wb.active

    header_font = Font(bold=True, size=14)
    for counter, field in enumerate(fields):
        if counter == 0:
            index = 0
        else:
            index = 1
        active_field_cell = ws.cell(row=ws.max_row, column=ws.max_column + index, value=field)
        active_field_cell.font = header_font

    for d in data:
        ws.append(d)
    return wb


class ExportFactory:
    def __init__(self, export_type, module, request):
        self.export_type = export_type
        self.module = module
        self.module_maps = {
            'JOB_TITLE': ExportJobTitle(request),
            'BANKS': ExportBankLists(request),
            'JOB_GRADES': ExportJobGrade(request),
            'NATURE_OF_CONTRACT': ExportNatureOfContract(request),
            # 'LEAVE_CATEGORY': ExportLeaveCategory(request),
            # 'LEAVE_TYPES': ExportLeaveTypes(request),
            'HOLIDAY': ExportHoliday(request),
            'EMPLOYEE': ExportEmployee(request)
        }
        self.request = request

    def get_module_class(self):
        module_class = self.module_maps[self.module]
        return module_class

    def prepare_export_data(self):
        return self.get_module_class().generate_export_data()

    def prepare_export_fields(self):
        fields_array = self.get_module_class().get_export_fields()
        formatted_fields = [name['display'] for name in fields_array]
        return formatted_fields

    def generate_export_file(self):
        if self.export_type == "EXCEL":
            return generate_excel(self.prepare_export_fields(), self.prepare_export_data())
        # print export PDF
        return
