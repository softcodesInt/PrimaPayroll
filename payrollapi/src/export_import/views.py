from django.http import HttpResponse
from rest_framework.generics import GenericAPIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl import load_workbook

from employees.models import Employee
from accounts.utils import get_parent_company
from .serializers import ExportSerializer, ImportEmployeeSerializer
from .ExportFactory.base import ExportFactory, ExportEmployee


class ExportFileView(GenericAPIView):
    serializer_class = ExportSerializer
    permission_classes = (IsAuthenticated,)

    def post(self, request):
        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)

        data = serializer.validated_data
        export_instance = ExportFactory(export_type=data['export_type'],
                                        module=data['module'],
                                        request=request)
        if data['export_type'] == 'EXCEL':
            workbook = export_instance.generate_export_file()
            response = HttpResponse(content=save_virtual_workbook(workbook),
                                    content_type="application/ms-excel")
            response['Content-Disposition'] = 'attachment; filename={}.xlsx'.format(data['module'])
            return response


class ImportEmployeeView(GenericAPIView):
    serializer_class = ImportEmployeeSerializer
    permission_classes = (IsAuthenticated,)

    def get_database_field(self, key):
        employee_struct = ExportEmployee(self.request).get_export_fields()
        field = None
        for struct in employee_struct:
            if struct['display'] == key:
                if struct.get('import_field'):
                    field = struct['import_field']
                else:
                    field = struct['value']
                break

        return field

    def post(self, request):
        employee_file = request.FILES['employeeFile']
        wb = load_workbook(filename=employee_file.file)
        ws = wb.active
        header = [cell.value for cell in ws[1]]

        excel_data = []

        for count, row in enumerate(ws.rows):
            if count == 0:
                continue
            data = {}
            for key, cell in zip(header, row):
                data[self.get_database_field(key)] = cell.value

            serializer = self.serializer_class(data=data, context={'request': request})
            if not serializer.is_valid(raise_exception=False):
                error = serializer.errors
                error['row_number'] = [f'Error occured at row: {count}']
                return Response(error, status=400)

            data = serializer.validated_data
            data['main_company'] = get_parent_company(request)
            excel_data.append(data)

        employees = []
        for d in excel_data:
            employees.append(Employee(**d))

        Employee.objects.bulk_create(employees)
        return Response({})
