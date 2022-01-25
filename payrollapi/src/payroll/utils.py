import io
import urllib3
from payroll.models import PayrollElements, TaxRelief
from employees.models import Employee
from employees.serializers import EmployeeSerializer
from rest_framework.exceptions import ValidationError

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font


def generate_bank_letter(request, company):
    employees = Employee.objects.filter(main_company=company)
    wb = Workbook()
    ws = wb.active

    http = urllib3.PoolManager()
    r = http.request('GET', company.logo.url)
    image_file = io.BytesIO(r.data)
    img = Image(image_file)
    img.width = 260
    img.height = 100
    ws.add_image(img, 'B2')
    header_name = ws.cell(row=9, column=3, value=company.name)
    header_name.font = Font(bold=True, size=22)
    # fix time
    ws.cell(row=10, column=2, value="EMPLOYEE SALARY")
    ws.append([])

    # table starts here
    fields = ['Employee Code', 'Bank Name',  'Account Number', 'Account Holder Name', 'Net Pay']
    bold_font = Font(bold=True)
    employee_code_cell = ws.cell(row=12, column=1, value="Employee Code")
    employee_code_cell.font = bold_font

    account_holder_cell = ws.cell(row=12, column=2, value="Account Holder Name")
    account_holder_cell.font = bold_font

    bank_name_cell = ws.cell(row=12, column=3, value="Bank Name")
    bank_name_cell.font = bold_font

    account_number_cell = ws.cell(row=12, column=4, value="Account Number")
    account_number_cell.font = bold_font

    net_pay_cell = ws.cell(row=12, column=5, value="Net Pay")
    net_pay_cell.font = bold_font

    total_amount_to_pay = 0

    for employee in employees:
        if not employee.is_employee_entitled_to_salary:
            continue
        data = EmployeeSerializer(instance=employee, context={'request': request}).data
        net_pay = data['total_earnings'] - data['total_deductions']
        data = (employee.employee_code, employee.bank.name, employee.account_number, employee.account_name, net_pay)
        ws.append(data)

        total_amount_to_pay += net_pay

    grand_total_cell = ws.cell(row=ws.max_row + 2, column=4, value="Grand Total")
    grand_total_cell.font = Font(bold=True, size=16)

    total_amount_cell = ws.cell(row=ws.max_row, column=5, value=total_amount_to_pay)
    total_amount_cell.font = Font(bold=True, size=16)
    return wb


def generate_payroll_report(request, company):
    employees = Employee.objects.filter(main_company=company)
    wb = Workbook()
    ws = wb.active

    payroll_elements = PayrollElements.objects.filter(category__company_policy__company__in=[company])
    tax_reliefs = TaxRelief.objects.filter(relief_group__company_policy__company__in=[company])
    headers = [
        'Employee Code',
        'Employee First Name',
        'Employee Last Name',
        'Job Title',
        'Nature of Contract'
    ]

    for payroll_element in payroll_elements:
        headers.append(payroll_element.name)

    for tax_relief in tax_reliefs:
        headers.append(tax_relief.name)

    # SAME NAME as the names in get_payroll_data of EmployeeSerializer
    headers.extend(['Loan', 'Payee Tax', 'Pension'])

    # ADD THESE LATER
    headers.extend(['Total Earnings', 'Total Deductions', 'Net Pay'])
    ws.append(headers)

    for employee in employees:
        employee_data = EmployeeSerializer(instance=employee, context={'request': request}).data
        payroll_data = employee_data['payroll_data']
        data = [
            employee.employee_code,
            employee.first_name,
            employee.last_name,
            employee.job_title.name if employee.job_title else 'N/A',
            employee.nature_of_contract.name if employee.nature_of_contract else 'N/A',
        ]
        for header in headers[5:]:
            found_the_item = True
            for pd in payroll_data['earnings']:
                if pd.get('name') == header:
                    found_the_item = True
                    data.append(pd.get('amount') if pd.get('amount') != "None" else 0)
                    break
                else:
                    found_the_item = False

            if not found_the_item:
                # then it's deductions
                for pd in payroll_data['deductions']:
                    if pd.get('name') == header:
                        found_the_item = True
                        data.append(pd.get('amount') if pd.get('amount') != "None" else 0)
                        break
                    else:
                        found_the_item = False

            if not found_the_item and header not in ['Total Earnings', 'Total Deductions', 'Net Pay']:
                # this means the employee doesn't have that header value
                data.append(0)
        # add total columns
        data.extend(
            [
                employee_data['total_earnings'],
                employee_data['total_deductions'],
                employee_data['total_earnings'] - employee_data['total_deductions']
            ]
        )
        ws.append(data)
    return wb
